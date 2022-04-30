''' Importer for Sberbank broker - reports from XLS files exported from account operations:
    'Операции по счету' > 'Сделки' и 'Зачисления/списания'
    TODO: chcp 65001 & set PYTHONIOENCODING=utf-8
'''
#import xlrd
import datetime
import re
import os
#import csv
#from importlib.resources import open_text

from beancount.core.amount import D
from beancount.core import data
from beancount.core import flags
from beancount.core import account
from beancount.core import amount
from beancount.core import position
from beancount.ingest import importer
# from xlrd.biffh import XLRDError
# from xlrd.xldate import xldate_as_datetime

from openpyxl import Workbook, workbook
from openpyxl import load_workbook

import warnings
warnings.simplefilter("ignore")

NOCOST = position.CostSpec(None, None, None, None, None, None)

class Importer(importer.ImporterProtocol):
    '''An importer for Sberbanks XLS files'''

    def __init__(self, general_agreement_id,
                 account_root,
                 account_cash,
                 account_currencyexchange,
                 account_dividends,
                 account_interest,
                 account_fees, 
                 account_gains,
                 account_external,
                 account_repo = None,
                 balance = True):
        self.general_agreement_id = general_agreement_id
        self.account_root = account_root
        self.account_cash = account_cash
        self.account_currencyexchange = account_currencyexchange
        self.account_dividends = account_dividends
        self.account_interest = account_interest
        self.account_fees = account_fees
        self.account_gains = account_gains
        self.account_external = account_external
        self.balance = balance
        self.account_repo = account_repo if account_repo else account_fees

        # self.isindb = {}
        # self.isincur = {} # dictionary of isin code with corresponding asset base currencies
        # self.isinbond = {} # True if bond
        # self.isinfv = {} # Face value
        # self.exchanges = {
        #                     'РЦБ':self.account_cash,
        #                     'Вал. рынок':self.account_currencyexchange,
        #                     'Фонд. рынок':self.account_cash
        #                 }
        # self.cur = ['c Доллар США', 'c Евро']

    def fix_ticker(self, ticker):
        if ticker == '29011^':
            return 'OFZ29011'
        if ticker == 'ENPL':
            return 'ENPG'
        return ticker

    def check_sber(self, xlsfile, genid):
        ''' * Verify if file from Sberbank broker
        '''
        #workbook = xlrd.open_workbook(xlsfile, logfile=open(os.devnull, 'w'))
        workbook = load_workbook(xlsfile, read_only=True)
        sheet = workbook.active
        val = sheet['A2'].value
        workbook.close()
        # Check general agreement id
        if val != genid:
            return False
        return True

    def identify(self, file):
        ''' * Match if the filename is file from Sberbank
        '''
        # Check if it isn't LibreOffice lock file
        if re.match(r"\.~lock", os.path.basename(file.name)):
            return False
        # Match extension - should be XLSX
        if os.path.splitext(file.name)[1] != ".xlsx":
            return False
        # Check file name format and correct general agreement id
        if not (re.match(r"Сделки_", os.path.basename(file.name)) or re.match(r"Зачисления-и-Списания_", os.path.basename(file.name))):
            return False
        # Check if we have general agreement id in file
        return self.check_sber(file.name, self.general_agreement_id)

    def file_account(self, _):
        ''' *
        '''
        return self.account_root

    def file_date(self, file):
        ''' * Extract the statement date from the file
        '''
        # No report creation date - use file creation date instead
        return None

    def extract(self, file):
        ''' Open XLS file and create directives
        '''
        entries = []
        
        # extract report period
        per =  os.path.basename(file.name)
        pos1 = per.index('_') + 1 # start of period
        pos2 = pos1 + 12 # end of period
        self.stmt_begin = datetime.datetime.strptime(per[pos1:pos1+10], '%Y-%m-%d').date()
        self.stmt_end = datetime.datetime.strptime(per[pos2:pos2+10], '%Y-%m-%d').date()
        
        workbook = load_workbook(file.name, read_only=True)

        if False: #use broker report to extract balances
            entries += self.get_balance(workbook, file)

        if re.match(r"Сделки_", os.path.basename(file.name)):
            entries += self.get_trn(workbook, file)
        
        if re.match(r"Зачисления-и-Списания_", os.path.basename(file.name)):
            entries += self.get_cflow(workbook, file)

        workbook.close()

        return entries

    def get_balance(self, workbook, file):
        ''' * Parse broker report for end of period balances - cash and assets
            In: XLS sheet, file
            Out: list of transactions
        '''
        try:
            sheet = workbook.sheet_by_name('Динамика позиций')
        except XLRDError:
            return None # No balances sheet in file

        result = []
        ii = 0
        market = 0
        asset_type = 0
        #acc = ''
        while sheet.row(ii)[6].value != 'Стоимость всех позиций, руб.':
            if sheet.row(ii)[6].value[:5] == 'Актив':
                hh = sheet.row(ii) # header of the table
                jj = 6
                while True:
                    if hh[jj].value == 'хранения':
                        market = jj
                        break
                    jj += 1
            # check if it's not the next section's head
            if sheet.row(ii)[2].value == 'Валюта':
                asset_type = 1 # Currency
            elif sheet.row(ii)[2].value == 'Акции' or sheet.row(ii)[2].value == 'Прочее':
                asset_type = 2 # Stocks and ADRs
            if sheet.row(ii)[6].value:
                meta = data.new_metadata(file.name, ii)
                if asset_type == 1:
                    # line with currencies
                    ticker = sheet.row(ii)[6].value
                    acc = self.exchanges[sheet.row(ii)[market].value]
                    result.append(data.Balance(meta, 
                                    self.stmt_end + datetime.timedelta(days=1),
                                    acc,
                                    amount.Amount(D(str(sheet.row(ii)[15].value)), ticker),
                                    None, None))
                elif asset_type == 2:
                    # line with stocks
                    ticker = self.isindb[sheet.row(ii+1)[6].value]
                    account_inst = account.join(self.account_root, ticker)
                    amt = amount.Amount(D(str(sheet.row(ii)[15].value)), ticker)
                    result.append(data.Balance(meta, 
                                    self.stmt_end + datetime.timedelta(days=1),
                                    account_inst,
                                    amt,
                                    None, None))
                    # each stock adr ADR occupy 2 lines so skip additional line
                    ii += 1
            ii += 1

        return result

    def get_trn(self, workbook, file):
        ''' Parse Сделки for all assets transactions
        '''
        sheet = workbook['Сделки']
        hh = self.proc_header(sheet)
        result = []
    
        for i, row in enumerate(sheet.values):
            # skip header
            if i == 0:
                continue
            
            meta = data.new_metadata(file.name, i)
            sign = -1 if row[hh['Операция']] == 'Продажа' else 1
            ticker_cur = row[hh['Валюта']]
            ticker = self.fix_ticker(row[hh['Код финансового инструмента']])
            desc = row[hh['Номер сделки']]
            amt = amount.Amount(sign*D(str(row[hh['Количество']])), ticker)
            trn_cost = amount.Amount(-1*sign*D(str(row[hh['Сумма зачисления/списания']])), ticker_cur)
            commision = amount.Amount(D(str(
                                    row[hh['Комиссия торговой системы']] + row[hh['Комиссия банка']]
                                                        )), ticker_cur)

            if row[hh['Тип сделки']] == 'РЕПО':
                trn_date_2 = row[hh['Дата расчётов по 2-й части сделки РЕПО']].date()
                delta = amount.Amount(D(str(
                    row[hh['Сумма зачисления/списания']] - row[hh['Сумма зачисления/списания 2-й части сделки РЕПО']])), 
                    ticker_cur)
                
                txn = data.Transaction(
                                        meta, trn_date_2, self.FLAG, None, desc, data.EMPTY_SET, data.EMPTY_SET, 
                                        [
                                            data.Posting(self.account_cash, delta, None, None, None, None),
                                            data.Posting(self.account_fees, commision, None, None, None, None),
                                            data.Posting(self.account_repo, None, None, None, None, None)
                                        ])
            else:
                trn_date = row[hh['Дата расчётов']].date()
                account_inst = account.join(self.account_root, ticker)
            
                txn = self.create_tx(meta, sign, trn_date, ticker, desc, amt, trn_cost, commision, account_inst, self.account_gains)

            result.append(txn)
        return result
        
    def create_tx(self, meta, sign, trn_date, ticker, desc, amt, trn_cost, commision, account_inst, account_gains):
        pp = [
                    data.Posting(self.account_cash, trn_cost, None, None, None, None),
                    data.Posting(self.account_fees, commision, None, None, None, None),
                    data.Posting(account_inst, amt, NOCOST, None, None, None),
                    ]
        if sign == -1:
            pp.append(data.Posting(account_gains.format(ticker), None, None, None, None, None))
        txn = data.Transaction(meta, trn_date, self.FLAG, None, desc, data.EMPTY_SET, data.EMPTY_SET, pp)
                    
        return txn

    def get_cflow(self, workbook, file):
        ''' Parse broker export for all cash transactions except commissions - we got them from assets transactions
        '''

        sheet = workbook['Движение ДС']
        hh = self.proc_header(sheet)
        result = []

        for i, row in enumerate(sheet.values):
            # skip header
            if i == 0:
                continue
            
            oper = row[hh['Операция']]
            desc = row[hh['Содержание операции']]

            # if oper not in ['Ввод ДС', 'Вывод ДС', 'Списание налогов']:
            #     continue
            if not ((oper in ['Ввод ДС', 'Вывод ДС', 'Списание налогов']) or (desc == 'Оплата депозитарных услуг')):
                continue
            sign = -1 if oper in ['Списание налогов', 'Вывод ДС', 'Списание комиссии'] else 1
            meta = data.new_metadata(file.name, i)
            trn_date = row[hh['Дата исполнения поручения']].date()
            ticker_cur = row[hh['Валюта операции']]
            
            amt = amount.Amount(sign*D(str(row[hh['Сумма']])), ticker_cur)

            txn = data.Transaction(
                            meta, trn_date, self.FLAG, None, desc, data.EMPTY_SET, {trn_date}, [
                                data.Posting(self.account_cash, amt, None, None, None,
                                                None),
                                data.Posting(self.account_external, -amt, None, None, None,
                                                None),
                            ])
            result.append(txn)
        return result

    def proc_header(self, sheet):
        ''' process header of the table
            input: sheet from Sberbank's export of operations
            return: dictionary of column's headers with their indices
        '''
        dd = {}
        row = next(sheet.values)
        for i, col in enumerate(row):
            dd[col] = i
        return dd

    def get_cashflow(self, book, sheet, index, file):
        ''' Parse broker report for all cash operations (deposits, drawback, fees, dividends)
            In: XLS sheet, index of section title(1.1.)
            Out: list of transactions -- empty if there is none
        '''
        result = []
        acc = ''
        dedup = []
        # find beggining of next block
        ii = index + 1
        while ii<sheet.nrows-1:
            # check if it's not the next section's head
            if (re.match('^1\.3', sheet.row(ii)[1].value) or sheet.row(ii)[1].value=='3. Активы:' 
                    or re.match('^2\.1', sheet.row(ii)[1].value) or re.match('^2\.3', sheet.row(ii)[1].value)):
                break

            if sheet.row(ii)[1].value[:6] == '1.1.1.':
                acc_choice = self.account_cash
                acc = self.account_cash
            if sheet.row(ii)[1].value[:6] == '1.1.2.':
                acc_choice = self.account_currencyexchange
                acc = self.account_currencyexchange

            xfx = sheet.cell_xf_index(ii, 1)
            xf = book.xf_list[xfx]
            bgx = xf.background.pattern_colour_index
            if bgx != 24:
                ii +=1
                continue
            
            tt = sheet.row(ii-1)[1].value
            if tt[:11] == 'Валюта цены':
                trn_currency = fix_currency(tt[tt.find('=')+2:tt.find(',')])
            else: 
                trn_currency = fix_currency(tt)
            ii += 1 # skip table head
            while ii<sheet.nrows-1:
                if sheet.row(ii)[2].value == 'Итого:':
                    ii += 1
                    continue
                if sheet.row(ii)[1].value[:15] == 'Итого по валюте':
                    ii += 4
                    break

                trn_date = datetime.datetime.strptime(sheet.row(ii)[1].value, '%d.%m.%y').date()
                try:
                    acc = self.exchanges[sheet.row(ii)[12].value]
                except KeyError:
                    acc = acc_choice
                meta = data.new_metadata(file.name, ii)
                if sheet.row(ii)[2].value == 'Приход ДС':
                    amt = amount.Amount(D(str(sheet.row(ii)[6].value)), trn_currency)
                    txn = data.Transaction(
                        meta, trn_date, self.FLAG, None, sheet.row(ii)[2].value, data.EMPTY_SET, {trn_date}, [
                            data.Posting(acc, amt, None, None, None,
                                            None),
                            data.Posting(self.account_external, -amt, None, None, None,
                                            None),
                        ])
                    result.append(txn)
                elif sheet.row(ii)[2].value == 'Вывод ДС':
                    amt = amount.Amount(D(str(sheet.row(ii)[7].value)), trn_currency)
                    txn = data.Transaction(
                        meta, trn_date, self.FLAG, None, sheet.row(ii)[2].value, data.EMPTY_SET, {trn_date}, [
                            data.Posting(acc, -amt, None, None, None,
                                            None),
                            data.Posting(self.account_external, amt, None, None, None,
                                            None),
                        ])
                    result.append(txn)
                elif sheet.row(ii)[2].value == 'Переводы между площадками':
                    try:
                        acc1 = self.exchanges[sheet.row(ii)[11].value]
                        acc2 = self.exchanges[sheet.row(ii)[13].value]
                    except KeyError:
                        try:
                            acc1 = self.exchanges[sheet.row(ii)[10].value]
                            acc2 = self.exchanges[sheet.row(ii)[12].value]
                        except KeyError:
                            acc1 = self.exchanges[sheet.row(ii)[12].value]
                            acc2 = self.exchanges[sheet.row(ii)[14].value]

                    if sheet.row(ii)[6].value:
                        x = sheet.row(ii)[6].value
                        amt = amount.Amount(D(str(sheet.row(ii)[6].value)), trn_currency)
                    else:
                        x = -sheet.row(ii)[7].value
                        amt = amount.Amount(-D(str(sheet.row(ii)[7].value)), trn_currency)
                    
                    dd = [acc1, acc2, x] if x>0 else [acc2, acc1, -x]
                    try:
                        dedup.index(dd)
                        dedup.remove(dd)
                        ii += 1
                        continue
                    except ValueError:
                        dedup.append(dd)

                    txn = data.Transaction(
                        meta, trn_date, self.FLAG, None, sheet.row(ii)[2].value, data.EMPTY_SET, {trn_date}, [
                            data.Posting(acc1, amt, None, None, None,
                                            None),
                            data.Posting(acc2, -amt, None, None, None,
                                            None),
                        ])
                    result.append(txn)
                elif sheet.row(ii)[2].value == 'Дивиденды':
                    # unfortunately we don't have ticker info for dividends
                    amt = amount.Amount(D(str(sheet.row(ii)[6].value)), trn_currency)
                    txn = data.Transaction(
                        meta, trn_date, self.FLAG, None, sheet.row(ii)[2].value, data.EMPTY_SET, {trn_date}, [
                            data.Posting(acc, amt, None, None, None,
                                            None),
                            data.Posting(self.account_dividends, -amt, None, None, None,
                                            None),
                        ])
                    result.append(txn)
                elif sheet.row(ii)[2].value in ['Урегулирование сделок','Вознаграждение за обслуживание счета депо', 'Хранение ЦБ',
                                                'Вознаграждение компании', 'Quik','Оплата за вывод денежных средств', 
                                                'Комиссия за займы "овернайт ЦБ"', 'Урегулирование сделок по Айсберг-заявкам']:
                    # unfortunately we don't have ticker info for fees
                    amt = amount.Amount(D(str(sheet.row(ii)[7].value)), trn_currency)
                    txn = data.Transaction(
                        meta, trn_date, self.FLAG, None, sheet.row(ii)[2].value, data.EMPTY_SET, {trn_date}, [
                            data.Posting(acc, -amt, None, None, None,
                                            None),
                            data.Posting(self.account_fees, amt, None, None, None,
                                            None),
                        ])
                    result.append(txn)
                elif sheet.row(ii)[2].value in ['Займы "овернайт"', 'Проценты по займам "овернайт"', 'Проценты по займам "овернайт ЦБ"',
                                                'НКД от операций', 'НДФЛ', 'Подоходный налог']:
                    if sheet.row(ii)[7].value:
                        amt = amount.Amount(D(str(sheet.row(ii)[7].value)), trn_currency) 
                        txn = data.Transaction(
                            meta, trn_date, self.FLAG, None, sheet.row(ii)[2].value, data.EMPTY_SET, {trn_date}, [
                            data.Posting(acc, -amt, None, None, None,
                                            None),
                            data.Posting(self.account_interest, amt, None, None, None,
                                            None),
                        ])
                        result.append(txn)
                    if sheet.row(ii)[6].value:
                        amt = amount.Amount(-D(str(sheet.row(ii)[6].value)), trn_currency)
                        txn = data.Transaction(
                            meta, trn_date, self.FLAG, None, sheet.row(ii)[2].value, data.EMPTY_SET, {trn_date}, [
                                data.Posting(acc, -amt, None, None, None,
                                                None),
                                data.Posting(self.account_interest, amt, None, None, None,
                                                None),
                            ])
                        result.append(txn)

                ii += 1
                if sheet.row(ii)[1].value[:5] == 'Итого':
                    ii += 4 # skip notification after table
                    break
            # it's next section - lets find if this is fees
            if sheet.row(ii)[1].value in ['1.2. Займы "Овернайт":', '1.3. Удержанные сборы/штрафы (итоговые суммы):', 
                                        '1.2. Займы "Овернайт"/"Овернайт ГО":', '1.2. Займы:']:
                return result
        return result

    def get_transactions(self, book, sheet, index, file):
        # Cash and papers are described in different subsections index += 1
        entries = []
        acc = ''
        ii = index
        while ii<sheet.nrows-1:
            # find heading of table - using background color
            xfx = sheet.cell_xf_index(ii, 1)
            xf = book.xf_list[xfx]
            bgx = xf.background.pattern_colour_index
            if bgx != 24:
                ii +=1
                continue
            # Find section with stocks
            if sheet.row(ii-2)[1].value in ['Акция', 'АДР', 'Пай']:
                acc = self.account_cash
                ii += 1
                # Transactions table begins in row+3 (first ticker) and continues until blank line
                while sheet.row(ii)[1].value != '':
                    # read ticker
                    ticker = fix_ticker(sheet.row(ii)[1].value)
                    account_inst = account.join(self.account_root, ticker)
                    isin = sheet.row(ii)[7].value
                    title = sheet.row(ii)[8].value
                    ii += 1
                    #start to read transactions - pass to first transaction
                    while sheet.row(ii)[1].value[:5] != r'Итого':
                        trn_date = datetime.datetime.strptime(sheet.row(ii)[1].value, '%d.%m.%y').date()
                        trn_num = sheet.row(ii)[2].value #transaction id by broker
                        trn_currency = 'RUB' if sheet.row(ii)[11].value == 'Рубль' else sheet.row(ii)[11].value
                        # are we selling our buying? cols 4,5 - buying, cols 7,8 - selling
                        if sheet.row(ii)[4].value != '':
                            # we bought the ticker
                            # instantiate amount bought, 'currency' - ticker itself
                            meta = data.new_metadata(file.name, ii)
                            units_inst = amount.Amount(D(sheet.row(ii)[4].value), ticker) # amount of ticker bought
                            price = amount.Amount(-1*D(str(sheet.row(ii)[6].value)), trn_currency) # payment for transaction
                            cost = position.Cost(D(str(sheet.row(ii)[5].value)), trn_currency, None, None) # cost of single unit
                            #cost = amount.Amount(D(str(sheet.row(ii)[5].value)), trn_currency) # cost of single unit
                            #pos = position.Position(units_inst, cost)
                            txn = data.Transaction(
                                    meta, trn_date, self.FLAG, None, title, data.EMPTY_SET, {trn_num}, [
                                        data.Posting(acc, price, None, None, None, None),
                                        #data.Posting(self.account_fees, fees, None, None, None,
                                        #            None), # no fees in transaction description
                                        data.Posting(account_inst, units_inst, cost, None, None, None),
                                    ])
                            entries.append(txn)
                        elif sheet.row(ii)[7].value != '':
                            # we sold ticker
                            meta = data.new_metadata(file.name, ii)
                            units_inst = amount.Amount(-1*D(sheet.row(ii)[7].value), ticker) # amount of ticker sold
                            price = amount.Amount(D(str(sheet.row(ii)[9].value)), trn_currency) # payment for transaction
                            cost = amount.Amount(D(str(sheet.row(ii)[8].value)), trn_currency) # cost of single unit
                            account_gains = self.account_gains.format(ticker)
                            txn = data.Transaction(
                                    meta, trn_date, self.FLAG, None, title, data.EMPTY_SET, {trn_num}, [
                                        data.Posting(acc, price, None, None, None,
                                                        None),
                                        #data.Posting(self.account_fees, fees, None, None, None,
                                        #                None), # no fees associated with transaction
                                        data.Posting(account_inst, units_inst, NOCOST, cost, None,
                                                        None),
                                        data.Posting(account_gains, None, None, None, None,
                                                        None),
                                ])
                            entries.append(txn)
                        ii +=1
                    ii +=1
            # pass line with totals for current ticker
            elif sheet.row(ii-2)[1].value == 'Иностранная валюта':
                # Currency conversion
                acc = self.account_currencyexchange
                # First ticker in 3d line from subsection head
                ii += 1
                while sheet.row(ii)[1].value != '':
                    sold_currency = fix_currency(sheet.row(ii)[5].value)
                    bought_currency = fix_currency(sheet.row(ii)[8].value)
                    ticker = sheet.row(ii)[1].value
                    ii += 1
                    while sheet.row(ii)[1].value[:5] != r'Итого':
                        meta = data.new_metadata(file.name, ii)
                        trn_date = datetime.datetime.strptime(sheet.row(ii)[1].value, '%d.%m.%y').date()
                        trn_num = sheet.row(ii)[2].value #transaction id by broker
                        # check if we buy or sell
                        if sheet.row(ii)[4].value:
                            # we buy
                            conv_rate = amount.Amount(D(str(sheet.row(ii)[4].value)), 
                                                      bought_currency)
                            amount_bought = amount.Amount(D(sheet.row(ii)[5].value), sold_currency) # amount of ticker sold
                            price = amount.Amount(-1*D(str(sheet.row(ii)[6].value)), bought_currency) # payment for transaction,                            
                        else:
                            #we sell
                            conv_rate = amount.Amount(D(str(sheet.row(ii)[7].value)), 
                                                    bought_currency)
                            amount_bought = amount.Amount(-1*D(sheet.row(ii)[8].value), sold_currency) # amount of ticker sold
                            price = amount.Amount(D(str(sheet.row(ii)[9].value)), bought_currency) # payment for transaction
                        txn = data.Transaction(
                                    meta, trn_date, self.FLAG, None, ticker, data.EMPTY_SET, {trn_num}, [
                                        #data.Posting(acc, amount_bought, conv_rate, None, None, None),
                                        data.Posting(acc, amount_bought, None, conv_rate, None, None),
                                        data.Posting(acc, price, None, None, None, None),
                                    ])
                        entries.append(txn)
                        ii +=1
                    ii +=1
            elif sheet.row(ii-2)[1].value == 'Облигация':
                acc = self.account_cash
                ii += 1
                # Transactions table begins in row+3 (first ticker) and continues until blank line
                while sheet.row(ii)[1].value != '':
                    # read ticker
                    ticker = fix_ticker(sheet.row(ii)[1].value)
                    account_inst = account.join(self.account_root, ticker)
                    isin = sheet.row(ii)[7].value
                    title = sheet.row(ii)[8].value
                    ii += 1
                    #start to read transactions - pass to first transaction
                    while sheet.row(ii)[1].value[:5] != r'Итого':
                        trn_date = datetime.datetime.strptime(sheet.row(ii)[1].value, '%d.%m.%y').date()
                        trn_num = sheet.row(ii)[2].value #transaction id by broker
                        trn_currency = 'RUB' if sheet.row(ii)[13].value == 'Рубль' else sheet.row(ii)[13].value
                        # are we selling our buying? cols 4,5 - buying, cols 7,8 - selling
                        if sheet.row(ii)[4].value != '':
                            # we bought the ticker
                            # instantiate amount bought, 'currency' - ticker itself
                            meta = data.new_metadata(file.name, ii)
                            units_inst = amount.Amount(D(sheet.row(ii)[4].value), ticker) # amount of ticker bought
                            price = amount.Amount(-1*D(str(sheet.row(ii)[6].value)), trn_currency) # payment for transaction
                            cost = position.Cost(D(str(sheet.row(ii)[5].value))*10, trn_currency, None, None) # cost of single unit
                            txn = data.Transaction(
                                    meta, trn_date, self.FLAG, None, title, data.EMPTY_SET, {trn_num}, [
                                        data.Posting(acc, price, None, None, None, None),
                                        #data.Posting(self.account_fees, fees, None, None, None,
                                        #            None), # no fees in transaction description
                                        data.Posting(account_inst, units_inst, cost, None, None, None),
                                    ])
                            entries.append(txn)
                        elif sheet.row(ii)[8].value != '':
                            # we sold ticker
                            meta = data.new_metadata(file.name, ii)
                            units_inst = amount.Amount(-1*D(sheet.row(ii)[8].value), ticker) # amount of ticker sold
                            price = amount.Amount(D(str(sheet.row(ii)[10].value)), trn_currency) # payment for transaction
                            cost = amount.Amount(D(str(sheet.row(ii)[9].value))*10, trn_currency) # cost of single unit
                            account_gains = self.account_gains.format(ticker)
                            txn = data.Transaction(
                                    meta, trn_date, self.FLAG, None, title, data.EMPTY_SET, {trn_num}, [
                                        data.Posting(acc, price, None, None, None,
                                                        None),
                                        #data.Posting(self.account_fees, fees, None, None, None,
                                        #                None), # no fees associated with transaction
                                        data.Posting(account_inst, units_inst, NOCOST, cost, None,
                                                        None),
                                        data.Posting(account_gains, None, None, None, None,
                                                        None),
                                ])
                            entries.append(txn)
                        ii +=1
                    ii +=1
            
            ii +=1
    
        return entries
